#!/usr/bin/env python3
"""
Simple script to examine the Excel file structure
"""
import openpyxl
from datetime import datetime

def examine_excel_file(file_path):
    """Examine the Excel file and show its structure"""
    print("🔍 Examining Excel file:", file_path)
    print("=" * 50)
    
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print(f"📊 File info:")
    print(f"   Max rows: {ws.max_row}")
    print(f"   Max columns: {ws.max_column}")
    
    # Get headers
    print(f"\n📋 Column headers:")
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        headers.append(header)
        print(f"   Column {col}: {header}")
    
    # Show first 5 data rows
    print(f"\n📝 First 5 data rows:")
    for row_num in range(2, min(7, ws.max_row + 1)):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_num, column=col).value
            row_data.append(cell_value)
        print(f"   Row {row_num}: {row_data}")
    
    # Analyze data types and patterns
    print(f"\n🔍 Data analysis:")
    
    # Count unique plant names
    plant_names = set()
    dates = set()
    for row_num in range(2, ws.max_row + 1):
        # Plant name (column 2)
        plant_name = ws.cell(row=row_num, column=2).value
        if plant_name:
            plant_names.add(plant_name)
        
        # Date (column 1)  
        date_value = ws.cell(row=row_num, column=1).value
        if date_value:
            dates.add(date_value)
    
    print(f"   Unique plants: {len(plant_names)}")
    print(f"   Plant names: {sorted(list(plant_names))}")
    print(f"   Unique dates: {len(dates)}")
    
    # Show date range
    if dates:
        date_objects = []
        for d in dates:
            if isinstance(d, str):
                try:
                    date_obj = datetime.strptime(d, "%d.%m.%Y").date()
                    date_objects.append(date_obj)
                except:
                    pass
            elif hasattr(d, 'date'):
                date_objects.append(d.date())
        
        if date_objects:
            date_objects.sort()
            print(f"   Date range: {date_objects[0]} to {date_objects[-1]}")
    
    wb.close()
    return headers, len(plant_names), len(dates)

if __name__ == "__main__":
    examine_excel_file("data/blumen_data.xlsx") 