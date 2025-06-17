from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from dateutil.parser import parse
from ..models.plant import Plant
import os
import time

class ExcelHandler:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self._ensure_file_exists()
        
        # Cache attributes
        self._cache: Optional[List[Dict[str, Any]]] = None
        self._cache_timestamp: Optional[float] = None
        self._file_mtime: Optional[float] = None
        
        # Preload cache on initialization
        self._load_cache()
    
    def _load_cache(self):
        """Load data into cache"""
        try:
            self._cache = self._read_data_uncached()
            self._cache_timestamp = time.time()
            self._file_mtime = os.path.getmtime(self.file_path)
        except Exception as e:
            print(f"Error loading cache: {str(e)}")
            self._cache = None
    
    def _is_cache_valid(self) -> bool:
        """Check if cache is still valid"""
        if self._cache is None or self._file_mtime is None:
            return False
        
        try:
            current_mtime = os.path.getmtime(self.file_path)
            return current_mtime == self._file_mtime
        except:
            return False
    
    def _invalidate_cache(self):
        """Invalidate the cache"""
        self._cache = None
        self._cache_timestamp = None
        self._file_mtime = None
    
    def _ensure_file_exists(self):
        """Ensure the Excel file exists, create if it doesn't"""
        if not self.file_path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            # Add headers
            headers = ["date", "plant name", "days without water", "water", "fertilizer", "wash", "neemoil", "pestmix", "size"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(self.file_path)
    
    def _get_plant_names(self, ws) -> List[str]:
        """Get unique plant names from the Excel file, preserving original order"""
        plant_names = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[1] not in plant_names:  # plant name column
                plant_names.append(row[1])
        return plant_names
    
    def _ensure_dates_exist(self, ws):
        """Ensure entries exist for today and next 7 days, with empty row separators between date groups"""
        today = datetime.now().date()
        end_date = today + timedelta(days=7)
        
        # Step 1: Collect all date information and find last rows
        existing_dates = set()
        date_to_rows = {}  # Maps each date to all its row indices
        
        # Go through all rows to collect date information
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:  # Skip rows without a date
                continue
                
            try:
                date = parse(row[0]).date() if isinstance(row[0], str) else row[0].date()
                existing_dates.add(date)
                
                if date not in date_to_rows:
                    date_to_rows[date] = []
                date_to_rows[date].append(row_idx)
            except:
                pass
        
        # Step 2: Process each date to ensure it has a separator after its last plant
        for date, row_indices in date_to_rows.items():
            last_row_idx = max(row_indices)  # Last row for this date
            
            # Check if we need to add a separator after this date
            if last_row_idx + 1 <= ws.max_row:
                next_row = [ws.cell(row=last_row_idx + 1, column=col).value for col in range(1, ws.max_column + 1)]
                if any(next_row):  # Next row has content (not empty)
                    # Insert empty row as separator
                    ws.insert_rows(last_row_idx + 1)
            else:
                # At end of sheet, add separator
                ws.append([None] * ws.max_column)
        
        # Get plant names
        plant_names = self._get_plant_names(ws)
        
        # Step 3: Add missing dates
        current_date = today
        while current_date <= end_date:
            if current_date not in existing_dates:
                # Add entries for each plant
                for plant in plant_names:
                    ws.append([
                        current_date.strftime("%d.%m.%Y"),
                        plant,
                        "",  # days without water
                        "",  # water
                        "",  # fertilizer
                        "",  # wash
                        "",  # neemoil
                        "",  # pestmix
                        ""   # size
                    ])
                # Add separator row
                ws.append([None] * ws.max_column)
            
            current_date += timedelta(days=1)
    
    def _get_last_care_date(self, ws, plant_name: str, care_type: str) -> datetime:
        """Get the last date a plant received care based on the correct logic for water and fertilizer."""
        last_date = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == plant_name:
                # Parse date robustly
                try:
                    date = None
                    if isinstance(row[0], str):
                        try:
                            date = datetime.strptime(row[0], "%d.%m.%Y").date()
                        except ValueError:
                            date = parse(row[0]).date()
                    elif isinstance(row[0], datetime):
                        date = row[0].date()
                    if not date:
                        continue
                except Exception:
                    continue

                if care_type == "water":
                    days_wo_water = row[2]
                    water_entry = row[3]  # Index 3 is the "water" column
                    # Check both "days without water" and "water" columns
                    if days_wo_water == 0 or (isinstance(days_wo_water, str) and days_wo_water.strip() == "0") or water_entry:
                        last_date = date
                elif care_type == "fertilizer":
                    fertilizer_value = row[4]
                    if fertilizer_value and str(fertilizer_value).strip() != "":
                        last_date = date
        return last_date

    def get_todays_plants(self) -> List[Plant]:
        """Get all plants with their current care status using cached data"""
        try:
            # Use cached data instead of loading workbook
            data = self.read_data()  # This will use cache
            
            today = datetime.now().date()
            plants = []
            
            # Get unique plant names from cached data
            plant_names = []
            for row in data:
                plant_name = row.get("plant name")
                if plant_name and plant_name not in plant_names:
                    plant_names.append(plant_name)
            
            for idx, plant_name in enumerate(plant_names, 1):
                # Get last care dates from cached data
                last_watered = self._get_last_care_date_from_cache(data, plant_name, "water")
                last_fertilized = self._get_last_care_date_from_cache(data, plant_name, "fertilizer")
                
                # Calculate days since last care
                days_since_watering = (today - last_watered).days if last_watered else None
                days_since_fertilizing = (today - last_fertilized).days if last_fertilized else None
                
                # Create plant object
                plant = Plant(
                    id=idx,
                    name=plant_name,
                    last_watered=last_watered,
                    last_fertilized=last_fertilized,
                    days_since_watering=days_since_watering,
                    days_since_fertilizing=days_since_fertilizing,
                    watering_schedule=7,  # default to weekly
                    fertilizing_schedule=14,  # default to bi-weekly
                    needs_water=days_since_watering is not None and days_since_watering >= 7,
                    needs_fertilizer=days_since_fertilizing is not None and days_since_fertilizing >= 14
                )
                plants.append(plant)
            
            return plants
        except Exception as e:
            print(f"Error in get_todays_plants: {str(e)}")
            raise
    
    def read_data(self) -> List[Dict[str, Any]]:
        """Read data from Excel file with caching"""
        # Check if cache is valid
        if self._is_cache_valid() and self._cache is not None:
            return self._cache.copy()  # Return a copy to prevent external modifications
        
        # Cache is invalid or doesn't exist, reload
        self._load_cache()
        return self._cache.copy() if self._cache is not None else []
    
    def _read_data_uncached(self) -> List[Dict[str, Any]]:
        """Read data from Excel file without caching (internal use)"""
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = wb.active
        
        # Ensure dates exist
        self._ensure_dates_exist(ws)
        wb.save(self.file_path)
        
        # Read data
        data = []
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                data.append(dict(zip(headers, row)))
        
        return data
    
    def write_data(self, data: List[Dict[str, Any]]):
        """Write data to Excel file while preserving formatting"""
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        
        # Clear existing data (except headers)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        # Write new data
        headers = [cell.value for cell in ws[1]]
        for row_data in data:
            row = []
            for header in headers:
                row.append(row_data.get(header, ""))
            ws.append(row)
        
        wb.save(self.file_path)
        
        # Invalidate cache after writing
        self._invalidate_cache()
    
    def get_plant_history(self, plant_name: str) -> List[Dict[str, Any]]:
        """
        Get all historical entries for a specific plant, ordered by date.
        Uses cached data for better performance.
        """
        # Use cached data
        data = self.read_data()
        
        # Filter for this plant
        plant_data = []
        for row in data:
            if row.get("plant name") == plant_name and row.get("date"):
                # Parse date
                date_value = row["date"]
                if isinstance(date_value, str):
                    try:
                        date_value = datetime.strptime(date_value, "%d.%m.%Y").date()
                    except ValueError:
                        try:
                            date_value = parse(date_value).date()
                        except:
                            continue  # Skip if date can't be parsed
                elif isinstance(date_value, datetime):
                    date_value = date_value.date()
                else:
                    continue  # Skip if date is missing or invalid
                
                row_copy = row.copy()
                row_copy["date"] = date_value  # Replace with parsed date
                plant_data.append(row_copy)
        
        # Sort by date (oldest first)
        plant_data.sort(key=lambda x: x["date"])
        
        return plant_data
    
    def _get_last_care_date_from_cache(self, data: List[Dict[str, Any]], plant_name: str, care_type: str) -> Optional[datetime]:
        """Get the last date a plant received care from cached data"""
        last_date = None
        
        for row in data:
            if row.get("plant name") == plant_name:
                # Parse date robustly
                try:
                    date = None
                    date_value = row.get("date")
                    if isinstance(date_value, str):
                        try:
                            date = datetime.strptime(date_value, "%d.%m.%Y").date()
                        except ValueError:
                            date = parse(date_value).date()
                    elif isinstance(date_value, datetime):
                        date = date_value.date()
                    if not date:
                        continue
                except Exception:
                    continue

                if care_type == "water":
                    days_wo_water = row.get("days without water")
                    water_entry = row.get("water")  # Get the "water" column value
                    # Check both "days without water" and "water" columns
                    if days_wo_water == 0 or (isinstance(days_wo_water, str) and str(days_wo_water).strip() == "0") or water_entry:
                        last_date = date
                elif care_type == "fertilizer":
                    fertilizer_value = row.get("fertilizer")
                    if fertilizer_value and str(fertilizer_value).strip() != "":
                        last_date = date
        
        return last_date 