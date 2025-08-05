#!/usr/bin/env python3
"""
Verify Excel to PostgreSQL migration was successful (simplified for uv).
"""

import os
from pathlib import Path
from datetime import datetime, date
import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, func, text
from sqlalchemy.orm import sessionmaker

# Import our models
from create_tables_simple import Plant, DailyCare

# Load environment variables
load_dotenv()

def setup_database():
    """Setup database connection."""
    db_config = {
        'host': os.getenv('DB_HOST'),
        'port': os.getenv('DB_PORT', '5432'),
        'database': os.getenv('DB_NAME'),
        'username': os.getenv('DB_USER'),
        'password': os.getenv('DB_PASSWORD')
    }
    
    connection_string = (
        f"postgresql://{db_config['username']}:{db_config['password']}"
        f"@{db_config['host']}:{db_config['port']}/{db_config['database']}"
    )
    
    engine = create_engine(connection_string, echo=False)
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    return SessionLocal

def count_excel_data(excel_path: str) -> dict:
    """Count records in Excel for comparison."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    stats = {
        'total_rows': 0,
        'unique_plants': set(),
        'unique_dates': set(),
        'water_events': 0,
        'fertilizer_events': 0,
        'treatment_events': 0
    }
    
    for row_num in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row_num, column=1).value
        plant_name = ws.cell(row=row_num, column=2).value
        water = ws.cell(row=row_num, column=4).value
        fertilizer = ws.cell(row=row_num, column=5).value
        wash = ws.cell(row=row_num, column=7).value
        neemoil = ws.cell(row=row_num, column=8).value
        pestmix = ws.cell(row=row_num, column=9).value
        
        if date_val and plant_name:
            stats['total_rows'] += 1
            stats['unique_plants'].add(str(plant_name).strip())
            
            if isinstance(date_val, datetime):
                stats['unique_dates'].add(date_val.date())
            
            if water and str(water).strip():
                stats['water_events'] += 1
            
            if fertilizer and str(fertilizer).strip():
                stats['fertilizer_events'] += 1
            
            if any(x and str(x).strip() for x in [wash, neemoil, pestmix]):
                stats['treatment_events'] += 1
    
    wb.close()
    
    stats['unique_plants'] = len(stats['unique_plants'])
    stats['unique_dates'] = len(stats['unique_dates'])
    
    return stats

def count_database_data(session_factory) -> dict:
    """Count records in database."""
    stats = {}
    
    with session_factory() as session:
        stats['total_plants'] = session.query(Plant).count()
        stats['total_care_records'] = session.query(DailyCare).count()
        
        # Unique dates
        unique_dates = session.query(func.count(func.distinct(DailyCare.care_date))).scalar()
        stats['unique_dates'] = unique_dates
        
        # Activity counts
        stats['water_events'] = session.query(DailyCare).filter(DailyCare.water_ml.isnot(None)).count()
        stats['fertilizer_events'] = session.query(DailyCare).filter(DailyCare.fertilizer.isnot(None)).count()
        stats['treatment_events'] = session.query(DailyCare).filter(DailyCare.treatment.isnot(None)).count()
    
    return stats

def show_sample_data(session_factory):
    """Show sample migrated data."""
    print("\n🔍 Sample Migrated Data:")
    print("-" * 30)
    
    with session_factory() as session:
        # Show some plants
        plants = session.query(Plant).limit(5).all()
        print("Sample plants:")
        for plant in plants:
            print(f"  • ID {plant.id}: {plant.name}")
        
        # Show some care records
        care_records = session.query(DailyCare, Plant.name)\
            .join(Plant)\
            .limit(5)\
            .all()
        
        print("\nSample care records:")
        for care, plant_name in care_records:
            activities = []
            if care.water_ml:
                activities.append(f"💧{care.water_ml}ml")
            if care.fertilizer:
                activities.append(f"🌿{care.fertilizer}")
            if care.treatment:
                activities.append(f"🧴{care.treatment}")
            if care.condition:
                activities.append(f"📝{care.condition}")
            
            activity_str = " + ".join(activities) if activities else "No activities"
            print(f"  • {care.care_date} | {plant_name} | {activity_str}")

def test_relationships(session_factory):
    """Test that foreign key relationships work."""
    print("\n🔗 Testing Relationships:")
    print("-" * 25)
    
    with session_factory() as session:
        # Test join query
        join_count = session.query(DailyCare).join(Plant).count()
        care_count = session.query(DailyCare).count()
        
        if join_count == care_count:
            print("✅ All care records have valid plant references")
        else:
            print(f"❌ Found orphaned care records: {care_count - join_count}")
        
        # Test relationship navigation
        sample_plant = session.query(Plant).first()
        if sample_plant:
            care_count = len(sample_plant.care_records)
            print(f"✅ Relationship test: {sample_plant.name} has {care_count} care records")
        
        sample_care = session.query(DailyCare).first()
        if sample_care:
            plant_name = sample_care.plant.name
            print(f"✅ Back-reference test: Care record belongs to {plant_name}")

def show_interesting_queries(session_factory):
    """Show some interesting queries you can now run."""
    print("\n🎯 Interesting Database Queries:")
    print("-" * 35)
    
    with session_factory() as session:
        # Most recent watering
        recent_watering = session.query(Plant.name, DailyCare.care_date, DailyCare.water_ml)\
            .join(DailyCare)\
            .filter(DailyCare.water_ml.isnot(None))\
            .order_by(DailyCare.care_date.desc())\
            .first()
        
        if recent_watering:
            print(f"Most recent watering: {recent_watering[0]} got {recent_watering[2]}ml on {recent_watering[1]}")
        
        # Plant with most care records
        plant_with_most_care = session.query(Plant.name, func.count(DailyCare.id).label('care_count'))\
            .join(DailyCare)\
            .group_by(Plant.id, Plant.name)\
            .order_by(func.count(DailyCare.id).desc())\
            .first()
        
        if plant_with_most_care:
            print(f"Most tracked plant: {plant_with_most_care[0]} with {plant_with_most_care[1]} care records")
        
        # Total water used
        total_water = session.query(func.sum(DailyCare.water_ml)).scalar() or 0
        print(f"Total water used: {total_water}ml")

def main():
    """Run migration verification."""
    print("✅ Migration Verification")
    print("=" * 30)
    
    # Setup database
    try:
        session_factory = setup_database()
        print("✅ Database connection successful")
    except Exception as e:
        print(f"❌ Database connection failed: {e}")
        return False
    
    # Compare Excel vs Database
    print("\n📊 Comparing Excel vs Database:")
    
    excel_path = "data/blumen_data.xlsx"
    if not Path(excel_path).exists():
        print(f"❌ Excel file not found: {excel_path}")
        return False
    
    excel_stats = count_excel_data(excel_path)
    db_stats = count_database_data(session_factory)
    
    print(f"\n📋 Excel File:")
    print(f"   Total rows: {excel_stats['total_rows']}")
    print(f"   Unique plants: {excel_stats['unique_plants']}")
    print(f"   Unique dates: {excel_stats['unique_dates']}")
    print(f"   Water events: {excel_stats['water_events']}")
    print(f"   Fertilizer events: {excel_stats['fertilizer_events']}")
    print(f"   Treatment events: {excel_stats['treatment_events']}")
    
    print(f"\n🗄️  Database:")
    print(f"   Total plants: {db_stats['total_plants']}")
    print(f"   Total care records: {db_stats['total_care_records']}")
    print(f"   Unique dates: {db_stats['unique_dates']}")
    print(f"   Water events: {db_stats['water_events']}")
    print(f"   Fertilizer events: {db_stats['fertilizer_events']}")
    print(f"   Treatment events: {db_stats['treatment_events']}")
    
    # Analysis
    print(f"\n🔍 Analysis:")
    if excel_stats['unique_plants'] == db_stats['total_plants']:
        print("   ✅ Plant count matches perfectly")
    else:
        diff = excel_stats['unique_plants'] - db_stats['total_plants']
        print(f"   ⚠️  Plant count difference: {diff}")
    
    coverage = (db_stats['total_care_records'] / excel_stats['total_rows']) * 100 if excel_stats['total_rows'] > 0 else 0
    print(f"   📊 Migration coverage: {coverage:.1f}% of Excel rows")
    
    # Test relationships
    test_relationships(session_factory)
    
    # Show sample data
    show_sample_data(session_factory)
    
    # Show interesting queries
    show_interesting_queries(session_factory)
    
    print(f"\n🎉 Verification complete!")
    
    if db_stats['total_plants'] > 0 and db_stats['total_care_records'] > 0:
        print("✅ Migration appears successful!")
        print("\n💡 You can now:")
        print("   • Query your data with SQL")
        print("   • Build APIs with this normalized structure")
        print("   • Calculate 'days without water' dynamically")
        print("   • Add new features easily")
    else:
        print("❌ Migration may have issues")
    
    return True

if __name__ == "__main__":
    main()