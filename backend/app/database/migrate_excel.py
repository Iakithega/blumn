#!/usr/bin/env python3
"""
Migrate Excel plant care data to PostgreSQL database.

This script:
1. Reads your Excel file (data/blumen_data.xlsx)
2. Extracts unique plants → inserts into plants table
3. Transforms care data → inserts into daily_care table  
4. Handles data cleaning and validation
5. Provides migration statistics
"""

import sys
from pathlib import Path
from datetime import datetime, date
from typing import Dict, Any, List, Optional
import openpyxl
from sqlalchemy.exc import IntegrityError

# Add the project root to Python path
project_root = Path(__file__).parent.parent.parent.parent
sys.path.append(str(project_root))

from backend.app.database.connection import db_manager, test_connection
from backend.app.database.models import Plant, DailyCare


class ExcelMigrator:
    """
    Handles migration from Excel to normalized database.
    
    Transformation logic:
    - Excel row: date | plant_name | water | fertilizer | wash | neemoil | pestmix
    - Database: plants table + daily_care table (normalized)
    """
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.stats = {
            'plants_created': 0,
            'care_records_created': 0,
            'rows_processed': 0,
            'rows_skipped': 0,
            'errors': []
        }
        
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    def extract_plant_names(self) -> List[str]:
        """
        Extract unique plant names from Excel file.
        These will become records in the plants table.
        """
        print("🔍 Extracting unique plant names...")
        
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb.active
        
        plant_names = set()
        
        # Skip header row, iterate through data
        for row_num in range(2, ws.max_row + 1):
            plant_name = ws.cell(row=row_num, column=2).value  # Column B
            
            if plant_name and isinstance(plant_name, str):
                # Clean up plant name
                cleaned_name = plant_name.strip()
                if cleaned_name:
                    plant_names.add(cleaned_name)
        
        wb.close()
        
        plant_list = sorted(list(plant_names))
        print(f"   Found {len(plant_list)} unique plants:")
        for name in plant_list:
            print(f"   • {name}")
        
        return plant_list
    
    def create_plants(self, plant_names: List[str]) -> Dict[str, int]:
        """
        Create plant records in database.
        Returns mapping of plant_name → plant_id for later use.
        """
        print(f"\n🌱 Creating {len(plant_names)} plant records...")
        
        plant_id_map = {}
        
        with db_manager.get_session() as session:
            for plant_name in plant_names:
                try:
                    # Check if plant already exists
                    existing_plant = session.query(Plant).filter(Plant.name == plant_name).first()
                    
                    if existing_plant:
                        plant_id_map[plant_name] = existing_plant.id
                        print(f"   ✓ {plant_name} (already exists)")
                    else:
                        # Create new plant
                        new_plant = Plant(name=plant_name)
                        session.add(new_plant)
                        session.flush()  # Get the ID without committing
                        
                        plant_id_map[plant_name] = new_plant.id
                        self.stats['plants_created'] += 1
                        print(f"   ✓ {plant_name} (created)")
                        
                except Exception as e:
                    self.stats['errors'].append(f"Error creating plant '{plant_name}': {e}")
                    print(f"   ❌ {plant_name}: {e}")
            
            # Commit all plant creations
            session.commit()
        
        print(f"   📊 {self.stats['plants_created']} new plants created")
        return plant_id_map
    
    def parse_date(self, date_value) -> Optional[date]:
        """Parse date from Excel cell (handles multiple formats)."""
        if not date_value:
            return None
        
        try:
            if isinstance(date_value, datetime):
                return date_value.date()
            elif isinstance(date_value, date):
                return date_value
            elif isinstance(date_value, str):
                # Try different date formats
                for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                    try:
                        return datetime.strptime(date_value.strip(), fmt).date()
                    except ValueError:
                        continue
        except Exception:
            pass
        
        return None
    
    def combine_treatments(self, wash: Any, neemoil: Any, pestmix: Any) -> Optional[str]:
        """
        Combine wash, neemoil, pestmix into single treatment field.
        Returns the treatment type or None if no treatments.
        """
        treatments = []
        
        if wash and str(wash).strip():
            treatments.append(f"wash({str(wash).strip()})")
        
        if neemoil and str(neemoil).strip():
            treatments.append(f"neemoil({str(neemoil).strip()})")
        
        if pestmix and str(pestmix).strip():
            treatments.append(f"pestmix({str(pestmix).strip()})")
        
        return ", ".join(treatments) if treatments else None
    
    def migrate_care_data(self, plant_id_map: Dict[str, int]):
        """
        Migrate care data from Excel to daily_care table.
        This is the main transformation logic.
        """
        print(f"\n📅 Migrating care data...")
        
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb.active
        
        batch_size = 100
        batch_records = []
        
        with db_manager.get_session() as session:
            # Process each Excel row
            for row_num in range(2, ws.max_row + 1):
                try:
                    # Extract row data (based on your Excel structure)
                    row_data = {
                        'date': ws.cell(row=row_num, column=1).value,      # A: date
                        'plant_name': ws.cell(row=row_num, column=2).value, # B: plant name
                        'days_without_water': ws.cell(row=row_num, column=3).value, # C: calculated field (skip)
                        'water': ws.cell(row=row_num, column=4).value,      # D: water
                        'fertilizer': ws.cell(row=row_num, column=5).value, # E: fertilizer
                        'days_without_fertilizer': ws.cell(row=row_num, column=6).value, # F: calculated field (skip)
                        'wash': ws.cell(row=row_num, column=7).value,       # G: wash
                        'neemoil': ws.cell(row=row_num, column=8).value,    # H: neemoil
                        'pestmix': ws.cell(row=row_num, column=9).value,    # I: pestmix
                        'condition': ws.cell(row=row_num, column=11).value, # K: condition
                    }
                    
                    # Validate and clean data
                    care_date = self.parse_date(row_data['date'])
                    plant_name = row_data['plant_name']
                    
                    if not care_date or not plant_name:
                        self.stats['rows_skipped'] += 1
                        continue
                    
                    plant_name = str(plant_name).strip()
                    if plant_name not in plant_id_map:
                        self.stats['rows_skipped'] += 1
                        self.stats['errors'].append(f"Row {row_num}: Unknown plant '{plant_name}'")
                        continue
                    
                    # Transform care data
                    water_ml = None
                    if row_data['water'] is not None and str(row_data['water']).strip():
                        try:
                            water_ml = int(float(str(row_data['water']).strip()))
                        except (ValueError, TypeError):
                            pass
                    
                    fertilizer = None
                    if row_data['fertilizer'] is not None and str(row_data['fertilizer']).strip():
                        fertilizer = str(row_data['fertilizer']).strip()[:50]  # Limit length
                    
                    treatment = self.combine_treatments(
                        row_data['wash'], 
                        row_data['neemoil'], 
                        row_data['pestmix']
                    )
                    
                    condition = None
                    if row_data['condition'] is not None and str(row_data['condition']).strip():
                        condition = str(row_data['condition']).strip()
                    
                    # Create daily care record
                    care_record = DailyCare(
                        plant_id=plant_id_map[plant_name],
                        care_date=care_date,
                        water_ml=water_ml,
                        fertilizer=fertilizer,
                        treatment=treatment,
                        condition=condition
                    )
                    
                    batch_records.append(care_record)
                    self.stats['rows_processed'] += 1
                    
                    # Process in batches for better performance
                    if len(batch_records) >= batch_size:
                        session.add_all(batch_records)
                        try:
                            session.commit()
                            self.stats['care_records_created'] += len(batch_records)
                            print(f"   ✓ Processed {self.stats['care_records_created']} records...")
                        except IntegrityError as e:
                            session.rollback()
                            # Handle duplicate entries (plant_id + care_date must be unique)
                            print(f"   ⚠️  Batch had duplicates, processing individually...")
                            for record in batch_records:
                                try:
                                    session.add(record)
                                    session.commit()
                                    self.stats['care_records_created'] += 1
                                except IntegrityError:
                                    session.rollback()
                                    self.stats['rows_skipped'] += 1
                        
                        batch_records = []
                
                except Exception as e:
                    self.stats['errors'].append(f"Row {row_num}: {e}")
                    self.stats['rows_skipped'] += 1
                    continue
            
            # Process remaining records
            if batch_records:
                session.add_all(batch_records)
                try:
                    session.commit()
                    self.stats['care_records_created'] += len(batch_records)
                except IntegrityError:
                    session.rollback()
                    for record in batch_records:
                        try:
                            session.add(record)
                            session.commit()
                            self.stats['care_records_created'] += 1
                        except IntegrityError:
                            session.rollback()
                            self.stats['rows_skipped'] += 1
        
        wb.close()
        print(f"   📊 {self.stats['care_records_created']} care records created")
    
    def migrate(self):
        """Run the complete migration process."""
        print("🚀 Starting Excel → PostgreSQL Migration")
        print("=" * 50)
        
        try:
            # Step 1: Extract and create plants
            plant_names = self.extract_plant_names()
            plant_id_map = self.create_plants(plant_names)
            
            # Step 2: Migrate care data
            self.migrate_care_data(plant_id_map)
            
            # Step 3: Show results
            self.show_migration_summary()
            
        except Exception as e:
            print(f"❌ Migration failed: {e}")
            raise
    
    def show_migration_summary(self):
        """Display migration statistics."""
        print(f"\n📊 Migration Summary")
        print("=" * 30)
        print(f"✅ Plants created: {self.stats['plants_created']}")
        print(f"✅ Care records created: {self.stats['care_records_created']}")
        print(f"📋 Excel rows processed: {self.stats['rows_processed']}")
        print(f"⚠️  Rows skipped: {self.stats['rows_skipped']}")
        
        if self.stats['errors']:
            print(f"\n⚠️  Errors ({len(self.stats['errors'])}):")
            for error in self.stats['errors'][:10]:  # Show first 10 errors
                print(f"   • {error}")
            if len(self.stats['errors']) > 10:
                print(f"   ... and {len(self.stats['errors']) - 10} more")
        
        print(f"\n🎉 Migration completed successfully!")


def main():
    """Main migration function."""
    
    # Test database connection first
    print("🔍 Testing database connection...")
    if not test_connection():
        print("❌ Cannot connect to database!")
        sys.exit(1)
    
    # Run migration
    excel_file = "data/blumen_data.xlsx"
    migrator = ExcelMigrator(excel_file)
    migrator.migrate()
    
    print(f"\n💡 Next steps:")
    print(f"   1. Run: python query_examples.py")
    print(f"   2. Verify your data in the database")
    print(f"   3. Start building your plant care API!")


if __name__ == "__main__":
    main()