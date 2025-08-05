#!/usr/bin/env python3
"""
Migrate Excel data to PostgreSQL database (simplified for uv).

This script transforms your Excel structure to normalized database:
- Excel: One row per plant per day with all care info
- Database: Plants table + Daily_care table (normalized)
"""

import os
from pathlib import Path
from datetime import datetime, date
from typing import Dict, Any, List, Optional
import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import IntegrityError

# Import our models from the simple script
from create_tables_simple import Plant, DailyCare, Base

# Load environment variables
load_dotenv()

class ExcelMigrator:
    """Handles Excel to database migration."""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.stats = {
            'plants_created': 0,
            'care_records_created': 0,
            'rows_processed': 0,
            'rows_skipped': 0,
            'errors': []
        }
        
        # Setup database connection
        self._setup_database()
        
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    def _setup_database(self):
        """Setup database connection using .env credentials."""
        db_config = {
            'host': os.getenv('DB_HOST'),
            'port': os.getenv('DB_PORT', '5432'),
            'database': os.getenv('DB_NAME'),
            'username': os.getenv('DB_USER'),
            'password': os.getenv('DB_PASSWORD')
        }
        
        connection_string = (
            f"postgresql://{db_config['username']}:{db_config['password']}"
            f"@{db_config['host']}:{db_config['port']}/{db_config['database']}"
        )
        
        self.engine = create_engine(connection_string, echo=False)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
    
    def extract_plant_names(self) -> List[str]:
        """Extract unique plant names from Excel."""
        print("🔍 Extracting unique plant names from Excel...")
        
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb.active
        
        plant_names = set()
        
        # Process each row (skip header)
        for row_num in range(2, ws.max_row + 1):
            plant_name = ws.cell(row=row_num, column=2).value  # Column B: plant name
            
            if plant_name and isinstance(plant_name, str):
                cleaned_name = plant_name.strip()
                if cleaned_name:
                    plant_names.add(cleaned_name)
        
        wb.close()
        
        plant_list = sorted(list(plant_names))
        print(f"   Found {len(plant_list)} unique plants:")
        for name in plant_list:
            print(f"   • {name}")
        
        return plant_list
    
    def create_plants(self, plant_names: List[str]) -> Dict[str, int]:
        """Create plant records and return name→id mapping."""
        print(f"\n🌱 Creating {len(plant_names)} plant records...")
        
        plant_id_map = {}
        
        with self.SessionLocal() as session:
            for plant_name in plant_names:
                try:
                    # Check if plant exists
                    existing_plant = session.query(Plant).filter(Plant.name == plant_name).first()
                    
                    if existing_plant:
                        plant_id_map[plant_name] = existing_plant.id
                        print(f"   ✓ {plant_name} (already exists, ID: {existing_plant.id})")
                    else:
                        # Create new plant
                        new_plant = Plant(name=plant_name)
                        session.add(new_plant)
                        session.flush()  # Get ID without committing
                        
                        plant_id_map[plant_name] = new_plant.id
                        self.stats['plants_created'] += 1
                        print(f"   ✓ {plant_name} (created, ID: {new_plant.id})")
                
                except Exception as e:
                    self.stats['errors'].append(f"Error creating plant '{plant_name}': {e}")
                    print(f"   ❌ {plant_name}: {e}")
            
            session.commit()
        
        print(f"   📊 {self.stats['plants_created']} new plants created")
        return plant_id_map
    
    def parse_date(self, date_value) -> Optional[date]:
        """Parse date from Excel cell."""
        if not date_value:
            return None
        
        try:
            if isinstance(date_value, datetime):
                return date_value.date()
            elif isinstance(date_value, date):
                return date_value
            elif isinstance(date_value, str):
                # Try different date formats
                for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                    try:
                        return datetime.strptime(date_value.strip(), fmt).date()
                    except ValueError:
                        continue
        except Exception:
            pass
        
        return None
    
    def combine_treatments(self, wash: Any, neemoil: Any, pestmix: Any) -> Optional[str]:
        """Combine treatment columns into single field."""
        treatments = []
        
        if wash and str(wash).strip():
            treatments.append(f"wash({str(wash).strip()})")
        
        if neemoil and str(neemoil).strip():
            treatments.append(f"neemoil({str(neemoil).strip()})")
        
        if pestmix and str(pestmix).strip():
            treatments.append(f"pestmix({str(pestmix).strip()})")
        
        return ", ".join(treatments) if treatments else None
    
    def migrate_care_data(self, plant_id_map: Dict[str, int]):
        """Migrate care data from Excel to daily_care table."""
        print(f"\n📅 Migrating care data from Excel...")
        
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb.active
        
        batch_size = 100
        batch_records = []
        
        print(f"   Processing {ws.max_row - 1} Excel rows...")
        
        with self.SessionLocal() as session:
            # Process each Excel row
            for row_num in range(2, ws.max_row + 1):
                try:
                    # Extract Excel row data (based on your structure from examination)
                    row_data = {
                        'date': ws.cell(row=row_num, column=1).value,           # A: date
                        'plant_name': ws.cell(row=row_num, column=2).value,     # B: plant name
                        'days_without_water': ws.cell(row=row_num, column=3).value,  # C: calculated (skip)
                        'water': ws.cell(row=row_num, column=4).value,          # D: water
                        'fertilizer': ws.cell(row=row_num, column=5).value,     # E: fertilizer
                        'days_without_fertilizer': ws.cell(row=row_num, column=6).value,  # F: calculated (skip)
                        'wash': ws.cell(row=row_num, column=7).value,           # G: wash
                        'neemoil': ws.cell(row=row_num, column=8).value,        # H: neemoil
                        'pestmix': ws.cell(row=row_num, column=9).value,        # I: pestmix
                        'size': ws.cell(row=row_num, column=10).value,          # J: size (skip for now)
                        'condition': ws.cell(row=row_num, column=11).value,     # K: condition
                    }
                    
                    # Validate core data
                    care_date = self.parse_date(row_data['date'])
                    plant_name = row_data['plant_name']
                    
                    if not care_date or not plant_name:
                        self.stats['rows_skipped'] += 1
                        continue
                    
                    plant_name = str(plant_name).strip()
                    if plant_name not in plant_id_map:
                        self.stats['rows_skipped'] += 1
                        self.stats['errors'].append(f"Row {row_num}: Unknown plant '{plant_name}'")
                        continue
                    
                    # Transform care data
                    water_ml = None
                    if row_data['water'] is not None and str(row_data['water']).strip():
                        try:
                            water_ml = int(float(str(row_data['water']).strip()))
                        except (ValueError, TypeError):
                            pass
                    
                    fertilizer = None
                    if row_data['fertilizer'] is not None and str(row_data['fertilizer']).strip():
                        fertilizer = str(row_data['fertilizer']).strip()[:50]
                    
                    treatment = self.combine_treatments(
                        row_data['wash'], 
                        row_data['neemoil'], 
                        row_data['pestmix']
                    )
                    
                    condition = None
                    if row_data['condition'] is not None and str(row_data['condition']).strip():
                        condition = str(row_data['condition']).strip()
                    
                    # Create daily care record
                    care_record = DailyCare(
                        plant_id=plant_id_map[plant_name],
                        care_date=care_date,
                        water_ml=water_ml,
                        fertilizer=fertilizer,
                        treatment=treatment,
                        condition=condition
                    )
                    
                    batch_records.append(care_record)
                    self.stats['rows_processed'] += 1
                    
                    # Process in batches
                    if len(batch_records) >= batch_size:
                        self._process_batch(session, batch_records, row_num)
                        batch_records = []
                
                except Exception as e:
                    self.stats['errors'].append(f"Row {row_num}: {e}")
                    self.stats['rows_skipped'] += 1
                    continue
            
            # Process remaining records
            if batch_records:
                self._process_batch(session, batch_records, ws.max_row)
        
        wb.close()
        print(f"   📊 {self.stats['care_records_created']} care records created")
    
    def _process_batch(self, session, batch_records: List[DailyCare], current_row: int):
        """Process a batch of records with duplicate handling."""
        try:
            session.add_all(batch_records)
            session.commit()
            self.stats['care_records_created'] += len(batch_records)
            print(f"   ✓ Processed {self.stats['care_records_created']} records (row {current_row})...")
        except IntegrityError:
            # Handle duplicates by processing individually
            session.rollback()
            print(f"   ⚠️  Batch had duplicates, processing individually...")
            for record in batch_records:
                try:
                    session.add(record)
                    session.commit()
                    self.stats['care_records_created'] += 1
                except IntegrityError:
                    session.rollback()
                    self.stats['rows_skipped'] += 1
    
    def migrate(self):
        """Run complete migration process."""
        print("🚀 Starting Excel → PostgreSQL Migration")
        print("=" * 50)
        
        try:
            # Step 1: Extract and create plants
            plant_names = self.extract_plant_names()
            plant_id_map = self.create_plants(plant_names)
            
            # Step 2: Migrate care data
            self.migrate_care_data(plant_id_map)
            
            # Step 3: Show results
            self.show_summary()
            
        except Exception as e:
            print(f"❌ Migration failed: {e}")
            raise
    
    def show_summary(self):
        """Show migration summary."""
        print(f"\n📊 Migration Summary")
        print("=" * 30)
        print(f"✅ Plants created: {self.stats['plants_created']}")
        print(f"✅ Care records created: {self.stats['care_records_created']}")
        print(f"📋 Excel rows processed: {self.stats['rows_processed']}")
        print(f"⚠️  Rows skipped: {self.stats['rows_skipped']}")
        
        if self.stats['errors']:
            print(f"\n⚠️  Errors ({len(self.stats['errors'])}):")
            for error in self.stats['errors'][:5]:  # Show first 5
                print(f"   • {error}")
            if len(self.stats['errors']) > 5:
                print(f"   ... and {len(self.stats['errors']) - 5} more")
        
        print(f"\n🎉 Migration completed successfully!")
        
        # Show database stats
        with self.SessionLocal() as session:
            plant_count = session.query(Plant).count()
            care_count = session.query(DailyCare).count()
            print(f"\n📋 Database now contains:")
            print(f"   • {plant_count} plants")
            print(f"   • {care_count} care records")

def main():
    """Main migration function."""
    print("🌱 Blumn Plant Care - Excel Migration")
    print("=" * 40)
    
    excel_file = "data/blumen_data.xlsx"
    
    if not Path(excel_file).exists():
        print(f"❌ Excel file not found: {excel_file}")
        return False
    
    try:
        migrator = ExcelMigrator(excel_file)
        migrator.migrate()
        
        print(f"\n💡 Next steps:")
        print(f"   1. Run: uv run python backend/app/database/verify_migration_simple.py")
        print(f"   2. Explore your data with SQL queries!")
        print(f"   3. Build your plant care API!")
        
        return True
        
    except Exception as e:
        print(f"❌ Migration failed: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    main()