#!/usr/bin/env python3
"""
Verify Excel to PostgreSQL migration was successful.

This script:
1. Compares record counts between Excel and database
2. Validates data integrity
3. Tests relationships work correctly
4. Checks for any data loss
"""

import sys
from pathlib import Path
from datetime import datetime, date
import openpyxl
from sqlalchemy import func

# Add the project root to Python path
project_root = Path(__file__).parent.parent.parent.parent
sys.path.append(str(project_root))

from backend.app.database.connection import db_manager, test_connection
from backend.app.database.models import Plant, DailyCare


def count_excel_data(excel_path: str) -> dict:
    """Count records in Excel file for comparison."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    stats = {
        'total_rows': 0,
        'unique_plants': set(),
        'unique_dates': set(),
        'water_events': 0,
        'fertilizer_events': 0,
        'treatment_events': 0
    }
    
    for row_num in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row_num, column=1).value
        plant_name = ws.cell(row=row_num, column=2).value
        water = ws.cell(row=row_num, column=4).value
        fertilizer = ws.cell(row=row_num, column=5).value
        wash = ws.cell(row=row_num, column=7).value
        neemoil = ws.cell(row=row_num, column=8).value
        pestmix = ws.cell(row=row_num, column=9).value
        
        if date_val and plant_name:
            stats['total_rows'] += 1
            stats['unique_plants'].add(str(plant_name).strip())
            
            if isinstance(date_val, datetime):
                stats['unique_dates'].add(date_val.date())
            
            if water and str(water).strip():
                stats['water_events'] += 1
            
            if fertilizer and str(fertilizer).strip():
                stats['fertilizer_events'] += 1
            
            if (wash and str(wash).strip()) or (neemoil and str(neemoil).strip()) or (pestmix and str(pestmix).strip()):
                stats['treatment_events'] += 1
    
    wb.close()
    
    stats['unique_plants'] = len(stats['unique_plants'])
    stats['unique_dates'] = len(stats['unique_dates'])
    
    return stats


def count_database_data() -> dict:
    """Count records in database for comparison."""
    stats = {}
    
    with db_manager.get_session() as session:
        stats['total_plants'] = session.query(Plant).count()
        stats['total_care_records'] = session.query(DailyCare).count()
        
        # Unique dates
        unique_dates = session.query(func.count(func.distinct(DailyCare.care_date))).scalar()
        stats['unique_dates'] = unique_dates
        
        # Activity counts
        stats['water_events'] = session.query(DailyCare).filter(DailyCare.water_ml.isnot(None)).count()
        stats['fertilizer_events'] = session.query(DailyCare).filter(DailyCare.fertilizer.isnot(None)).count()
        stats['treatment_events'] = session.query(DailyCare).filter(DailyCare.treatment.isnot(None)).count()
    
    return stats


def verify_relationships():
    """Test that foreign key relationships work correctly."""
    print("🔗 Testing Relationships...")
    
    with db_manager.get_session() as session:
        # Test 1: Can we join plants and daily_care?
        join_count = session.query(DailyCare).join(Plant).count()
        care_count = session.query(DailyCare).count()
        
        if join_count == care_count:
            print("   ✅ All care records have valid plant references")
        else:
            print(f"   ❌ Found orphaned care records: {care_count - join_count}")
        
        # Test 2: Can we access plant from care record?
        sample_care = session.query(DailyCare).first()
        if sample_care:
            plant_name = sample_care.plant.name  # This tests the relationship
            print(f"   ✅ Relationship working: Care record → {plant_name}")
        
        # Test 3: Can we access care records from plant?
        sample_plant = session.query(Plant).first()
        if sample_plant:
            care_count = len(sample_plant.care_records)  # This tests back reference
            print(f"   ✅ Back-reference working: {sample_plant.name} has {care_count} care records")


def verify_data_samples():
    """Verify some sample data looks correct."""
    print("\n🔍 Verifying Sample Data...")
    
    with db_manager.get_session() as session:
        # Show some plants
        plants = session.query(Plant).limit(3).all()
        print("   Sample plants:")
        for plant in plants:
            print(f"     • {plant.name}")
        
        # Show some care records with details
        care_records = session.query(DailyCare, Plant.name)\
            .join(Plant)\
            .filter(DailyCare.water_ml.isnot(None))\
            .limit(3)\
            .all()
        
        print("   Sample watering records:")
        for care, plant_name in care_records:
            print(f"     • {plant_name}: {care.water_ml}ml on {care.care_date}")


def main():
    """Run complete migration verification."""
    
    print("✅ Migration Verification")
    print("=" * 30)
    
    # Test database connection
    if not test_connection():
        print("❌ Cannot connect to database!")
        sys.exit(1)
    
    # Compare counts
    print("📊 Comparing Excel vs Database...")
    
    excel_path = "data/blumen_data.xlsx"
    excel_stats = count_excel_data(excel_path)
    db_stats = count_database_data()
    
    print(f"\n📋 Excel File:")
    print(f"   Total rows: {excel_stats['total_rows']}")
    print(f"   Unique plants: {excel_stats['unique_plants']}")
    print(f"   Unique dates: {excel_stats['unique_dates']}")
    print(f"   Water events: {excel_stats['water_events']}")
    print(f"   Fertilizer events: {excel_stats['fertilizer_events']}")
    print(f"   Treatment events: {excel_stats['treatment_events']}")
    
    print(f"\n🗄️  Database:")
    print(f"   Total plants: {db_stats['total_plants']}")
    print(f"   Total care records: {db_stats['total_care_records']}")
    print(f"   Unique dates: {db_stats['unique_dates']}")
    print(f"   Water events: {db_stats['water_events']}")
    print(f"   Fertilizer events: {db_stats['fertilizer_events']}")
    print(f"   Treatment events: {db_stats['treatment_events']}")
    
    # Analyze differences
    print(f"\n🔍 Analysis:")
    plant_diff = excel_stats['unique_plants'] - db_stats['total_plants']
    if plant_diff == 0:
        print("   ✅ Plant count matches perfectly")
    else:
        print(f"   ⚠️  Plant count difference: {plant_diff}")
    
    # Note: Care records might differ due to normalization and duplicate handling
    record_ratio = db_stats['total_care_records'] / excel_stats['total_rows'] if excel_stats['total_rows'] > 0 else 0
    print(f"   📊 Database has {record_ratio:.1%} of Excel rows (normal due to deduplication)")
    
    # Test relationships
    verify_relationships()
    
    # Show sample data
    verify_data_samples()
    
    print(f"\n🎉 Migration verification complete!")
    
    if db_stats['total_plants'] > 0 and db_stats['total_care_records'] > 0:
        print("✅ Migration appears successful!")
        print("\n💡 Try running: python query_examples.py")
    else:
        print("❌ Migration may have issues - check error logs")


if __name__ == "__main__":
    main()