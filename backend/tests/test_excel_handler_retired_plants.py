import unittest
from datetime import datetime
from unittest.mock import patch

from openpyxl import Workbook

from backend.app.core.excel_handler import ExcelHandler, is_retired_plant


class RetiredPlantRegistryTest(unittest.TestCase):
    def test_marks_hoya_mathilde_as_retired(self):
        self.assertTrue(is_retired_plant("Hoya Mathilde"))

    def test_keeps_hoya_sabah_current(self):
        self.assertFalse(is_retired_plant("Hoya Sabah"))

    def test_normalizes_case_and_extra_spaces(self):
        self.assertTrue(is_retired_plant("  hoya   mathilde  "))


class ActivePlantListTest(unittest.TestCase):
    def test_get_plant_names_skips_retired_plants(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["date", "plant name"])
        ws.append(["20.05.2026", "Hoya Mathilde"])
        ws.append(["20.05.2026", "Hoya Sabah"])

        handler = ExcelHandler.__new__(ExcelHandler)
        plant_names = handler._get_plant_names(ws)

        self.assertEqual(plant_names, ["Hoya Sabah"])


class DateGenerationTest(unittest.TestCase):
    def test_german_historical_dates_do_not_hide_missing_august_dates(self):
        class FixedDatetime(datetime):
            @classmethod
            def now(cls, tz=None):
                return cls(2026, 8, 3)

        wb = Workbook()
        ws = wb.active
        ws.append([
            "date",
            "plant name",
            "days without water",
            "water",
            "fertilizer",
            "wash",
            "neemoil",
            "pestmix",
            "size",
        ])
        ws.append(["08.05.2026", "Hoya Sabah"])
        ws.append(["08.06.2026", "Hoya Sabah"])
        ws.append(["08.07.2026", "Hoya Sabah"])

        handler = ExcelHandler.__new__(ExcelHandler)
        with patch("backend.app.core.excel_handler.datetime", FixedDatetime):
            handler._ensure_dates_exist(ws)

        generated_dates = {
            row[0]
            for row in ws.iter_rows(min_row=2, values_only=True)
            if isinstance(row[0], str)
        }

        self.assertIn("05.08.2026", generated_dates)
        self.assertIn("06.08.2026", generated_dates)
        self.assertIn("07.08.2026", generated_dates)


if __name__ == "__main__":
    unittest.main()
